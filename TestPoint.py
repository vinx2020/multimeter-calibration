# -*- coding: utf-8 -*-
import os
import re #Regular expression operations
import time
import datetime
import tkinter as tk
import tkinter.messagebox
from tkinter import filedialog
from tkinter import ttk
from PIL import Image, ImageTk  #jpg
from openpyxl import load_workbook
import threading
from threading import Thread, Event
import pyvisa as visa
from lxml import etree
rm = visa.ResourceManager()
rg1 = rm.list_resources()
rg2 = list(rg1)
sem = threading.Semaphore()
a1 = list()
b1 = list()
ws = 0
wb = 0
inst_1 = 0
inst_2 = 0
data_c2 = 0
thread = 0
shablon = r'C:\\Users\\vgatto\\cernbox\\02_Python\\TestPoint4'


def pribor():
    lb.insert('end', '____________________________________________')
    lb.insert('end', 'Detected devices and ports:')
    v = len(rg1)
    i = -1
    while i < v - 1:
        i = i + 1
        lb.insert('end', rg2[i])
        combo1.configure(values=rg2)
        combo2.configure(values=rg2)
    lb.insert('end', '____________________________________________')

def config():
    global pool_config
    pool_config = m.get()[0:4]
    tree = etree.parse('C:\\Users\\vgatto\\cernbox\\02_Python\\TestPoint4\\' + pool_config +'\\config.xml')
    lb.insert('end', 'Please connect the following equipment :')
    for instrument in tree.findall('instrument'):
        nom = instrument.find('nom').text
        typ = instrument.find('typ').text
        lb.insert('end', '- ' + nom + ' ' +typ)

def connect_dut():
    global a1
    global inst_1
    inst_1 = rm.open_resource(combo1.get())
    data_1 = inst_1.query("*IDN?")
    a.set(data_1)
    a1 = data_1
    a1 = re.findall(r'\w+', a1)
    if a1[2] in ('2024B'): a10.set('Oscilloscope ' + a1[2] + ' ' + 'connected')
    if a1[4] in ('2000', '2001'): a10.set('Keithley ' + a1[4] + ' ' + 'connected')
    lb.insert('end', a10.get())
    lb.see('end')


def connect_inst2():
    global b1
    global inst_2
    if combo2.get()[:4] == 'ASRL':
        inst_2 = rm.open_resource(g.get(), baud_rate=9600, data_bits=8, write_termination='\r', read_termination='\r')
    if combo2.get()[:4] == 'GPIB':
        inst_2 = rm.open_resource(g.get())
    data_2 = inst_2.query("*IDN?")
    b.set(data_2)
    b1 = data_2
    b1 = re.findall(r'\w+', b1)
    b14.set('Calibrator ' + b1[0] + ' ' + b1[1] + ' ' + 'connected')
    lb.insert('end', b14.get())
    lb.see('end')

#Stop		
def arret(): #non fonctionnel
    sem.acquire()
    inst_2.write('*RST')
    inst_1.write('*RST')
    lb.insert('end', 'Test interrompu')
		
class call(Thread):
    def __init__(self, name_test, write_calib, write_DUT, read_DUT, Bandwidth, time):
        Thread.__init__(self)
        self.name_test = name_test
        self.write_calib = write_calib
        self.write_DUT = write_DUT
        self.read_DUT = read_DUT
        self.Bandwidth = Bandwidth
        self.time = time
        self.start()


    def run(self):
        sem.acquire()
        inst_1.write(self.write_DUT)
        inst_1.write(self.Bandwidth)
        time.sleep(1)
        inst_2.write('*CLS')
        inst_2.write(self.write_calib)
        inst_2.write('OPER')
        time.sleep(5)
        if a1[2] == '2024B':
            inst_1.write('MEASU:MEAS1:SOU CH1;TYP MEAN;VAL?')
        elif a1[4] == '2000':
            inst_1.write('READ?')
        time_1 = float(self.time)
        time.sleep(time_1)
        valeur_lue = float(inst_1.read())
        ws[self.read_DUT] = valeur_lue	
        ws['A18'] = m.get() #poolN
        ws['B18'] = a1[5] #serial
        ws['C18'] = a1[6] #firm
        ws['G18'] = float(h.get()) #temp
        ws['H18'] = float(k.get()) #Humid
        #Extraction dans excel des tolerances
        Low_tol = re.sub(r'[0-9]+', '', self.read_DUT) + '13'         #cellule contenant tolerance basse
        High_tol = re.sub(r'[0-9]+', '', self.read_DUT) + '14'        #cellule contenant tolerance haute
        data_L = float(ws[Low_tol].value)
        data_H = float(ws[High_tol].value)
        if data_L <= valeur_lue <= data_H: verdict = 'pass'
        else: verdict = 'fail'
        if a1[2] == '2024B': pooln='0373'
        elif a1[4] == '2000': pooln='0113'
        wb.save(shablon + '\\' + pooln + '\\' + pooln + '.xlsx')
        ws['E18'] = n.get() #technician
        ws['D18'] = e #date
        inst_2.write('STBY')
        lb.insert('end', self.name_test + ' : ' + str(data_L)+ ' | ' + str(valeur_lue) + ' | ' + str(data_H) + ' --> ' + verdict)
        time.sleep(1)
        progress1.step(1)
        sem.release()


class reset(Thread):
    def __init__(self):
        Thread.__init__(self)
        self.start()

    def run(self):
        sem.acquire()
        time.sleep(2)
        inst_2.write('*RST')
        inst_1.write('*RST')
        time.sleep(2)
        #progress1.step(1)
        sem.release()


class message(Thread):
    def __init__(self, text):
        Thread.__init__(self)
        self.text = text
        self.start()

    def run(self):
        sem.acquire()
        tk.messagebox.showinfo('Please', self.text)
        #progress1.step(1)
        sem.release()


def illustration(message, image):
    sem.acquire()
    top = tk.Toplevel(interface)
    x = interface.winfo_x()
    y = interface.winfo_y()
    top.geometry("600x300+%d+%d" % (x + 100, y + 200))
    image = Image.open('C:\\Users\\vgatto\\cernbox\\02_Python\\TestPoint4\\' + pool_config +'\\' + image)
    photo = ImageTk.PhotoImage(image)
    autor = ttk.Label(top, justify='left', text=message).place(x=10,y=5)
    mick = ttk.Label(top, image=photo).place(x=10,y=30)
    But = ttk.Button(top, text='OK', width=10, command=top.destroy).place(x=80,y=250)  
    top.transient(interface)
    top.grab_set()
    interface.wait_window(top)
    sem.release()


class cap(Thread):
    def __init__(self):
        Thread.__init__(self)
        self.start()

    def run(self):
        sem.acquire()
        global data_c2
        inst_1.write('CONF:CAP')
        time.sleep(5)
        inst_1.write('READ?')
        time.sleep(5)
        data_c1 = inst_1.read()
        data_c2 = float(data_c1)
        time.sleep(1)
        #progress1.step(1)
        sem.release()

class tek(Thread):
    def __init__(self):
        Thread.__init__(self)
        self.start()

    def run(self):
        sem.acquire()
        global data_c2
        inst_1.write('CALibrate:INTERNAL')
        progress1.step(1)
        sem.release()


def start():
    global ws
    global wb
    global thread
    enable_animation()
    thread = reset()
        
    if a1[2] == '2024B':
	#sub_tek_tps2000_1yr_dcv_bw_trig_samp_ver_rs232_o_cal
        wb = load_workbook(shablon + '\\' + pool_config + '\\0373.xlsx')
        ws = wb.active
        progress1.configure(maximum = 72)
        lb.insert('end', 'Start time:' + ' ' + date1)
        # Self test
        #thread = tek()
        thread = message('Self calibration is proceeding : press Ok when finished')
        thread = message('Connect calibrator output to UUT CH1')
        # DC Voltage gain accuracy
        thread = call('Ch.1 5mV/div', 'SCOPE VOLT; OUT 17.5 mV, 0 Hz', 'CH1:PRO 1;SCA 5E-3;ACQ:MOD AVE;NUMAV 16MEASU:MEAS1:SOU CH1;TYP MEAN', 'L18', '', '7')
        thread = call('Ch.1 5mV/div', 'SCOPE VOLT; OUT -17.5 mV, 0 Hz', '', 'M18', '', '7')
        thread = call('Ch.1 200mV/div', 'SCOPE VOLT; OUT 700 mV, 0 Hz', 'CH1:SCA 200E-3', 'N18', '', '7')
        thread = call('Ch.1 200mV/div', 'SCOPE VOLT; OUT -700 mV, 0 Hz', '', 'O18', '', '7')
        thread = call('Ch.1 2V/div', 'SCOPE VOLT; OUT 7 V, 0 Hz', 'CH1:SCA 2', 'P18', '', '7')
        thread = call('Ch.1 2V/div', 'SCOPE VOLT; OUT -7 V, 0 Hz', '', 'Q18', '', '7')
        thread = reset()
        thread = message('Calibration complete')		

    if a1[4] == '2000':
        wb = load_workbook(shablon + '\\' + pool_config + '\\0113.xlsx')
        ws = wb.active
        ws.insert_rows(18)
        progress1.configure(maximum = 37)
        lb.insert('end', 'Start time:' + ' ' + date1)
        #e.g. call('ac 0.1V', 'OUT 0.1 V, 20 Hz', 'CONF:VOLT:AC 0.1', 'P18', 'VOLT:AC:DET:BAND 3', '8', '99.99', '100.01')
        lb.insert('end', 'Voltage dc Precision')
        #thread = message('Connect DC voltage leads')
        thread = illustration('Connect DC voltage leads', '55xx_20xx_vz_2w.jpg')
        thread = call('dc 0.1V', 'OUT 0.1 V', 'CONF:VOLT:DC 0.1', 'L18', '', '3')
        thread = call('dc -0.1V', 'OUT -0.1 V', 'CONF:VOLT:DC 0.1', 'M18', '', '3')
        thread = call('dc 1V', 'OUT 1.0 V', 'CONF:VOLT:DC 1.0', 'N18', '', '3')
        thread = call('dc -1V', 'OUT -1.0 V', 'CONF:VOLT:DC 1.0', 'O18', '', '3')
        thread = call('dc 10.0', 'OUT 10.0 V', 'CONF:VOLT:DC 10', 'P18', '', '3')
        thread = call('dc -10.0', 'OUT -10.0 V', 'CONF:VOLT:DC 10', 'Q18', '', '3')
        thread = call('dc 100V', 'OUT 100.0 V', 'CONF:VOLT:DC 100', 'R18', '', '3')
        thread = call('dc -100V', 'OUT -100.0 V', 'CONF:VOLT:DC 100', 'S18', '', '3')   
        # thread = call('dc 1000V', 'OUT 1000.0 V', 'CONF:VOLT:DC 1000', 'T18', '', '3')
        # thread = call('dc -1000V', 'OUT -1000.0 V', 'CONF:VOLT:DC 1000', 'U18', '', '3')
        thread = reset()
        # VAC
        thread = call('Ac 0.1V 1kHz', 'OUT 0.1 V, 1 kHz', 'CONF:VOLT:AC 0.1', 'V18', 'VOLT:AC:DET:BAND 3', '3')
        thread = call('Ac 0.1V 40 kHz', 'OUT 0.1 V, 40 kHz', 'CONF:VOLT:AC 0.1', 'W18', 'VOLT:AC:DET:BAND 3', '3')
        thread = call('Ac 1.0V 1kHz', 'OUT 1.0 V, 1 kHz', 'CONF:VOLT:AC 1.0', 'X18', 'VOLT:AC:DET:BAND 3', '3')
        thread = call('Ac 1.0V 40 kHz', 'OUT 1.0 V, 40 kHz', 'CONF:VOLT:AC 1.0', 'Y18', 'VOLT:AC:DET:BAND 3', '3')
        thread = call('Ac 10.0V 1kHz', 'OUT 10.0 V, 1 kHz', 'CONF:VOLT:AC 10.0', 'Z18', 'VOLT:AC:DET:BAND 3', '3')
        thread = call('Ac 10.0V 40 kHz', 'OUT 10.0 V, 40 kHz', 'CONF:VOLT:AC 10.0', 'AA18', 'VOLT:AC:DET:BAND 3', '3')       
        thread = call('Ac 100.0V 1kHz', 'OUT 100.0 V, 1 kHz', 'CONF:VOLT:AC 100.0', 'AB18', 'VOLT:AC:DET:BAND 3', '8')
        thread = call('Ac 100.0V 40 kHz', 'OUT 100.0 V, 40 kHz', 'CONF:VOLT:AC 100.0', 'AC18', 'VOLT:AC:DET:BAND 3', '8')
        thread = call('Ac 700.0V 1kHz', 'OUT 700.0 V, 1 kHz', 'CONF:VOLT:AC 750.0', 'AD18', 'VOLT:AC:DET:BAND 3', '8')
        thread = call('Ac 320.0V 40 kHz', 'OUT 320.0 V, 40 kHz', 'CONF:VOLT:AC 750.0', 'AE18', 'VOLT:AC:DET:BAND 3', '8')
        thread = reset()
        # Frequency
        thread = call('fr 50.0Hz', 'OUT 0.1 V, 50 Hz', 'CONF:FREQ 0.1', 'AF18', 'FREQ:DET:BAND 20', '5')
        thread = call('fr 5kHz', 'OUT 0.1 V, 5 kHz', 'CONF:FREQ 0.1', 'AG18', 'FREQ:DET:BAND 20', '5')
        thread = call('fr 250 kHz', 'OUT 0.1 V, 250 kHz', 'CONF:FREQ 0.1', 'AH18', 'FREQ:DET:BAND 20', '5')
        thread = call('fr 450kHz', 'OUT 0.1 V, 450 kHz', 'CONF:FREQ 0.1', 'AI18', 'FREQ:DET:BAND 20', '5')
        # Capacitance
        # thread = message('Measuring capacitance. Pull the red wire out of the calibrator to compensate for the wires')
        # thread = cap()
        # thread = message('Put the wire back in place')
        # thread = call('cap 1 NF', 'OUT 1 NF', 'CONF:CAP 1 NF', 'AJ18', '', '5')
        # thread = call('cap 10 NF', 'OUT 10 NF', 'CONF:CAP 10 NF', 'AK18', '', '5')
        # thread = call('cap 100 NF', 'OUT 100 NF', 'CONF:CAP 100 NF', 'AL18', '', '5')
        # thread = call('cap 1000 NF', 'OUT 1 UF', 'CONF:CAP 1 UF', 'AM18', '', '5')
        # DCI
        thread = message('Switch wires to measure current')
        thread = reset()
        thread = call('dci 0.01A', 'OUT 0.01 A', 'CONF:CURR:DC 0.01', 'AJ18', '', '5')
        thread = call('dci -0.01A', 'OUT -0.01 A', 'CONF:CURR:DC 0.01', 'AK18', '', '5')
        thread = call('dci 0.1A', 'OUT 0.1 A', 'CONF:CURR:DC 0.1', 'AL18', '', '5')
        thread = call('dci 0.1A', 'OUT -0.1 A', 'CONF:CURR:DC 0.1', 'AM18', '', '5')
        thread = call('dci 1.0A', 'OUT 1.0 A', 'CONF:CURR:DC 1.0', 'AN18', '', '5')
        thread = call('dci -1.0 A', 'OUT -1.0 A', 'CONF:CURR:DC 1.0', 'AO18', '', '5')
        # thread = call('dci 2.9A', 'OUT 2.9 A', 'CONF:CURR:DC 3.0', 'AP18', '', '5')
        # thread = call('dci -2.9A', 'OUT -2.9 A', 'CONF:CURR:DC 3.0', 'AQ18', '', '5')
        # ACI
        thread = call('aci 1.0 @1kHz', 'OUT 1 A, 1 kHz', 'CONF:CURR:AC 1.0', 'AR18', 'CURR:AC:DET:BAND 20', '5')		
        thread = call('aci 2.19999A @1kHz', 'OUT 2.19999 A, 1 kHz', 'CONF:CURR:AC 3.0', 'AS18', 'CURR:AC:DET:BAND 20', '5')	
        # Ohm 4wire
        thread = message('Switch wires in four-wire to measure resistance')
        thread = reset()	
        thread = call('res4 100', 'OUT 100 OHM; ZCOMP WIRE4', 'CONF:FRES 100', 'AT18', '', '5')
        thread = call('res4 1000', 'OUT 1 KOHM; ZCOMP WIRE4', 'CONF:FRES 1000', 'AU18', '', '5')
        thread = call('res4 10000', 'OUT 10 KOHM; ZCOMP WIRE4', 'CONF:FRES 10000', 'AV18', '', '5')
        thread = call('res4 100000', 'OUT 100 KOHM; ZCOMP WIRE4', 'CONF:FRES 100000', 'AW18', '', '5')
        thread = message('unplug sense wire from calibrator and put them to normal output (sense first) ')
        thread = call('res4 1000000', 'OUT 1 MOHM; ZCOMP WIRE4', 'CONF:FRES 1000000', 'AX18', '', '5')
        thread = message('Switch wires in 2-wire to measure resistance')
        thread = reset()
		# Ohm 2wire
        thread = call('res2 10 MOHM', 'OUT 10 MOHM; ZCOMP WIRE2', 'CONF:RES 10000000', 'AY18', '', '5')
        thread = call('res2 100 MOHM', 'OUT 100 MOHM; ZCOMP WIRE2', 'CONF:RES 100000000', 'AZ18', '', '5')
        thread = message('Calibration complete')
        thread = reset()	


def about_win():
    top = tk.Toplevel(interface)
    top.title('About the program')
    top.iconbitmap(r'C:\Users\vgatto\cernbox\02_Python\TestPoint4\DMM\icon\icon.ico')
    top.resizable(0, 0)
    w = top.winfo_screenwidth()
    h = top.winfo_screenheight()
    w = w // 3
    h = h // 2
    w = w - 200
    h = h - 200
    top.geometry('270x225+{}+{}'.format(w, h))
    text1 = ('Calibrator test bench\rAutor: adapted from ITL\rSupported Instruments :\r Multimeters:Keithley 2000\r Oscilloscope:TDS2024B')
    autor = ttk.Label(top, justify='left', text=text1).place(x=60,y=5)
    But = ttk.Button(top, text='OK', width=10, command=top.destroy).place(x=80,y=180)  
    top.transient(interface)
    top.grab_set()
    interface.wait_window(top)


class AnimatedGif(object):
    def __init__(self, image_file_path):
        self._frames = []

        frame_num = 0
        while True:
            try:
                frame = tk.PhotoImage(file=image_file_path,
                                   format='gif -index {}'.format(frame_num))
            except tk.TclError:
                break
            self._frames.append(frame)
            frame_num += 1

    def __len__(self):
        return len(self._frames)

    def __getitem__(self, frame_num):
        return self._frames[frame_num]


def update_label_image(label, img4, ms_delay, frame_num):
    global cancel_id
    label.configure(image=img4[frame_num])
    frame_num = (frame_num+1) % len(img4)
    cancel_id = interface.after(
        ms_delay, update_label_image, label, img4, ms_delay, frame_num)

def enable_animation():
    global cancel_id
    if cancel_id is None:
        ms_delay = 1000 // len(img4)
        cancel_id = interface.after(
            ms_delay, update_label_image, animation, img4, ms_delay, 0)

def cancel_animation():
    global cancel_id
    if cancel_id is not None:
        interface.after_cancel(cancel_id)
        cancel_id = None


#Interface graphique
interface = tkinter.Tk()
interface.title('TestPoint 1.2')
interface.geometry('900x700')
interface.iconbitmap(r'C:\Users\vgatto\cernbox\02_Python\TestPoint4\icon\icon.ico')
interface.configure(background='#ECE5F0')
frame = tk.Frame(interface)
frame.grid()

#Style
ttk.Style().configure('TButton', padding=6, font='arial 10', foreground='black', background='#ECE5F0')
ttk.Style().configure('BW.Label', padding=6, font='arial 10', foreground='black', background='#ECE5F0')
ttk.Style().configure('TLabelframe', background='#ECE5F0')
ttk.Style().configure("TProgressbar", foreground='blue', background='blue')

#Menu
main_menu = tk.Menu(interface)
file_menu = tk.Menu(main_menu, tearoff=False)
file_menu.add_command(label='New')
file_menu.add_command(label='Open')
file_menu.add_command(label='Save')
file_menu.add_separator()
file_menu.add_command(label='Close', command=interface.quit)
main_menu.add_cascade(label='File', menu=file_menu)
main_menu.add_cascade(label='Settings')
main_menu.add_cascade(label='About the program', command=about_win)

#Variables
today = datetime.datetime.today()
a = tk.StringVar()
b = tk.StringVar()
d = today.strftime('%d.%m.%Y.%H.%M.%S')
date1 = today.strftime('%H:%M:%S')
e = today.strftime('%d/%m/%Y')
f = tk.StringVar()
g = tk.StringVar()
h = tk.StringVar()
k = tk.StringVar()
l = tk.StringVar()
m = tk.StringVar()
n = tk.StringVar()
a10 = tk.StringVar()
b10 = tk.StringVar()
b14 = tk.StringVar()
b15 = tk.StringVar()
b16 = tk.StringVar()
img1 = tk.PhotoImage(file=r'C:\Users\vgatto\cernbox\02_Python\TestPoint4\icon\pan2.gif')
img2 = tk.PhotoImage(file=r'C:\Users\vgatto\cernbox\02_Python\TestPoint4\icon\start1.gif')
img3 = tk.PhotoImage(file=r'C:\Users\vgatto\cernbox\02_Python\TestPoint4\icon\ref1.gif')
img4 = AnimatedGif(r'C:\Users\vgatto\cernbox\02_Python\TestPoint4\icon\progress.gif')
img5 = tk.PhotoImage(file=r'C:\Users\vgatto\cernbox\02_Python\TestPoint4\icon\stop.gif')
cancel_id = None

#Entete
label = tk.Label(interface, image=img1)
label.place(x=1, y=1)

#Calibration conditions
lbf2 = ttk.LabelFrame(interface, text='Calibration conditions', width=410, height=80, style='TLabelframe')
lbf2.place(x=5, y=63)

lab7 = ttk.Label(interface, text='Pool n.:', style='BW.Label')
lab7.place(x=15, y=85)
entry7 = ttk.Entry(interface, textvariable=m, width=8, font='arial 8')
entry7.place(x=15, y=110)
lab8 = ttk.Label(interface, text='Operator:', style='BW.Label')
lab8.place(x=90, y=85)
entry8 = ttk.Entry(interface, textvariable=n, width=20, font='arial 8')
entry8.place(x=90, y=110)
temp = ttk.Label(interface, text='°C:', style='BW.Label')
temp.place(x=250, y=85)
entry4 = ttk.Entry(interface, textvariable=h, width=4, font='arial 8')
entry4.place(x=250, y=110)
lab5 = ttk.Label(interface, text='%R:', style='BW.Label')
lab5.place(x=300, y=85)
entry5 = ttk.Entry(interface, textvariable=k, width=4, font='arial 8')
entry5.place(x=300, y=110)
but_conf = ttk.Button(interface, text='Ok', width=4, command=config, style='TButton')
but_conf.place(x=350, y=100)


#Instruments
lbf1 = ttk.LabelFrame(interface, text='Identification', width=410, height=405, style='TLabelframe')
lbf1.place(x=5, y=148)

lab1 = ttk.Label(interface, text='ID DUT:', style='BW.Label')
lab1.place(x=15, y=175)
combo1 = ttk.Combobox(interface, state='readonly', height=4, width=35)
combo1.place(x=65, y=175)
but1 = ttk.Button(interface, text='Connect', width=10, command=connect_dut, style='TButton')
but1.place(x=310, y=175)
entry1 = ttk.Entry(interface, textvariable=a, state='readonly', width=38, font='arial 8')
entry1.place(x=65, y=200)

lab2 = ttk.Label(interface, text='ID Inst2:', style='BW.Label')
lab2.place(x=15, y=235)
combo2 = ttk.Combobox(interface, textvariable=g, state='readonly', height=4, width=35)
combo2.place(x=65, y=235)
but2 = ttk.Button(interface, text='Connect', width=10, command=connect_inst2, style='TButton')
but2.place(x=310, y=235)
entry2 = ttk.Entry(interface, textvariable=b, state='readonly', width=38, font='arial 8')
entry2.place(x=65, y=260)


#Actions
lbf3 = ttk.LabelFrame(interface, text='Progress', width=675, height=50, style='TLabelframe')
lbf3.place(x=5, y=625)

progress1 = ttk.Progressbar(interface, orient='horizontal', mode='determinate', length = 640, value = 0, style='TProgressbar')
progress1.place(x=15, y=645)
animation = ttk.Label(interface,image=img4[0], style='TLabel')
animation.place(x=695, y=632)
but4 = ttk.Button(interface, image=img3, command=pribor, style='TButton')
but4.place(x=725, y=632)
but3 = ttk.Button(interface, image=img2, command=start, style='TButton')
but3.place(x=775, y=632)
but_stop = ttk.Button(interface, image=img5, command=arret, style='TButton')
but_stop.place(x=825, y=632)


#Screen
lb = tk.Listbox(interface, selectmode='extended', width=75, height=34, relief='ridge')
lb.place(x=432, y=70)

interface.event_add('<<Paste>>', '<Control-igrave>')
interface.event_add("<<Copy>>", "<Control-ntilde>")
interface.config(menu=main_menu)

pribor()
interface.mainloop()